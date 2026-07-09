"""product_mapping.py — 상품 매핑/제외 리졸버 (옵션 기반 집계 + 사은품 제외 + 세트 분해)
~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
- 매핑표 xlsx → scripts/product_mapping.json 설정 생성(build_config)  [1회/매핑 변경 시]
- resolve(상품명, 옵션) → 정리된 상품명 리스트(마사지기만; 사은품·세트 비마사지기 구성품 제외)
  · 리스트가 비면([]) = 순수 사은품 리뷰 → 집계 제외 대상
  · 세트는 구성품(마사지기)들로 분해 → 리뷰 1건이 여러 상품에 귀속

우선순위:  ① 규칙(원본명+옵션 정확일치)  ② 유의어(별칭 부분일치, 긴 것 우선)  ③ 원본명 그대로
그 후: 세트면 구성품 분해 → 각 구성품을 다시 정규화 → 비마사지기 제외
"""
import json, re, os, sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
CFG_PATH = Path(__file__).resolve().parent / "product_mapping.json"

# 세트 정리된명 → 마사지기 구성품(정리된명) 목록. (10종, 명시적 — 파싱보다 안전)
# 비마사지기(마그네슘/파우치 등)는 애초에 넣지 않음.
# 담당자 확정 보정 (매핑표 내부 불일치 해소) — 별칭(소문자) → canonical
CANON_OVERRIDE = {
    "손편한케어 프리미엄 손 마사지기": "손편한케어",       # #5 통일
    "코어 요추벨트 허리 보호대": "코어 요추벨트",
}
CANON_FOOT = {"발편한케어 V1", "발편한케어 V2", "종아리편한케어"}


def _ver_of(opt):
    o = str(opt or "")
    if "V2" in o or "프로" in o:
        return "V2"
    if "V1" in o or "데일리" in o:
        return "V1"
    return None


SET_COMPONENTS = {
    "데일리 릴렉스 세트 (팔꿈치 마사지기 + 마그네슘 시너지 크림)": ["프리미엄 엘보케어 팔꿈치 마사지기"],
    "핸드 밸런스 세트 (팔꿈치 마사지기 + 손편한케어)": ["프리미엄 엘보케어 팔꿈치 마사지기", "손편한케어 프리미엄 손 마사지기"],
    "시그니처 세트 (허리편한케어 V2 + 목편한케어 플라잉)": ["허리편한케어 V2", "목편한케어 플라잉"],
    "코어 밸런스 업 세트 (허리편한케어 V2 + 목 마사지 베개 V2)": ["허리편한케어 V2", "목 마사지 베개 V2"],
    "바디 밸런스 케어 세트 (골반 마사지기 + 허리베개) + 마그네슘 시너지 크림 증정": ["골반 마사지기", "허리베개"],
    "목마사지베개 V2 + 마그네슘 시너지 크림 30ml": ["목 마사지 베개 V2"],
    "목마사지베개V2 + 마그네슘 시너지 크림 30ml": ["목 마사지 베개 V2"],
    "허리펀한케어V2 + 목마사지베개V2 + 마그네슘 시너지 크림 100ml": ["허리편한케어 V2", "목 마사지 베개 V2"],
    "★김마통SET★ 목편한케어 플라잉 + 허리편한케어V1 + 마그네슘 시너지 크림 100ml": ["목편한케어 플라잉", "허리편한케어 V1"],
    "★덤순이SET★ 넥숄더 힐링케어V2 + 허리편한케어V2 + 코어요추벨트": ["넥숄더 힐링케어 V2", "허리편한케어 V2", "코어 요추벨트 허리 보호대"],
}


def build_config(xlsx_path: str):
    """매핑표 xlsx + 제외 xlsx → product_mapping.json 생성."""
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    rules = {}
    for r in list(wb["매핑표(규칙적용)"].iter_rows(values_only=True))[1:]:
        if r[2] is not None:
            rules[str(r[2]).strip() + "" + str(r[3] or "").strip()] = str(r[0]).strip()
    aliases = []
    canon = set()
    for r in list(wb["요약"].iter_rows(values_only=True))[1:]:
        if not r[0]:
            continue
        clean = str(r[0]).strip()
        canon.add(clean)
        aliases.append([clean.lower(), clean])
        if r[1]:
            for a in str(r[1]).split(","):
                a = a.strip()
                if a:
                    aliases.append([a.lower(), clean])
    aliases.sort(key=lambda x: -len(x[0]))

    # 제외(비마사지기) — 그룹명 + 유의어 전부 소문자 별칭
    excl_path = Path(xlsx_path).parent / "리뷰 대시보드 리뷰 제외 품목.xlsx"
    excl_aliases = []
    if excl_path.is_file():
        wx = openpyxl.load_workbook(str(excl_path), data_only=True)
        for r in list(wx["비마사지기_상품리스트"].iter_rows(values_only=True))[1:]:
            grp = str(r[0]).strip() if r[0] else ""
            if grp:
                excl_aliases.append(grp.lower())
            if r[1]:
                for a in str(r[1]).split(","):
                    a = a.strip()
                    if a:
                        excl_aliases.append(a.lower())
    excl_aliases = sorted(set(excl_aliases), key=lambda x: -len(x))

    cfg = {"rules": rules, "aliases": aliases, "exclude_aliases": excl_aliases,
           "set_components": SET_COMPONENTS, "canonical": sorted(canon)}
    CFG_PATH.write_text(json.dumps(cfg, ensure_ascii=False, indent=2), encoding="utf-8")
    return cfg


class ProductResolver:
    def __init__(self, cfg=None):
        if cfg is None:
            cfg = json.loads(CFG_PATH.read_text(encoding="utf-8"))
        self.rules = cfg["rules"]
        self.aliases = cfg["aliases"]
        self.excl = cfg["exclude_aliases"]
        self.sets = cfg.get("set_components", {})
        # 정확일치 별칭 → canonical (요약 시트 기준). 규칙시트가 원본명을 그대로 둔 경우까지 병합.
        self._canon = {}
        for a, clean in self.aliases:
            self._canon.setdefault(a, clean)

    def canon(self, name):
        """전체 문자열 정확일치로 canonical 병합 (골반케어프리미엄→골반 마사지기 등). 없으면 원본."""
        lo = str(name).strip().lower()
        if lo in CANON_OVERRIDE:            # 담당자 확정 보정 우선
            return CANON_OVERRIDE[lo]
        return self._canon.get(lo, str(name).strip())

    def _foot_fix(self, clean, name, opt):
        """넥숄더 힐링케어/발편한케어/EMS/종아리/44언니 프로모 잔여 보정. None=해당없음, list=확정(빈=제외)."""
        n = str(name or ""); o = str(opt or "")
        # 넥숄더 힐링케어 버전: V2 명시 없으면 V1 ('넥숄더 프로'는 미해당)
        if "넥숄더 힐링케어" in clean or "넥숄더 힐링케어" in n:
            v = "V2" if ("V2" in clean or "V2" in n or "V2" in o) else "V1"
            return ["넥숄더 힐링케어 " + v]
        if clean in CANON_FOOT:             # 이미 정상 canonical → 통과
            return None
        # 44언니 프로모 페이지: '제품 선택'으로 실제 상품 결정
        if "44언니" in n or "제품 선택" in o:
            picks = []
            if "EMS 슬리퍼" in o or ("EMS" in o and "슬리퍼" in o):
                picks.append("발편한케어 V2")
            if ("발V2" in o) or ("발" in o and "V2" in o and "종아리" in o):
                picks.append("발편한케어 V2")
            if "종아리" in o:
                picks.append("종아리편한케어")
            return list(dict.fromkeys(picks))   # 빈 리스트 → 애매, 제외
        # 발편한케어 계열(EMS 발 포함) → 옵션 버전으로, 없으면 V1
        if ("발편한케어" in clean) or ("발편한케어" in n) or ("EMS" in n and "발" in n):
            return ["발편한케어 " + (_ver_of(o) or "V1")]
        # 종아리 단독
        if "종아리" in clean or "종아리" in n:
            return ["종아리편한케어"]
        return None

    def _is_excluded_name(self, clean):
        lo = str(clean).lower()
        return any(a in lo or lo in a for a in self.excl)

    def _clean_name(self, name, opt):
        key = str(name).strip() + "" + str(opt or "").strip()
        if key in self.rules:
            return self.rules[key]
        hay = (str(opt or "") + " " + str(name)).lower()
        for a, clean in self.aliases:
            if len(a) >= 3 and a in hay:
                return clean
        return str(name).strip()

    def resolve_ex(self, name, opt):
        """→ (마사지기 정리된명 리스트, is_set). 빈 리스트=집계제외. is_set=세트상품 리뷰 여부."""
        # 규칙 우선(세트 판정 정확), 규칙 없으면 별칭/원본
        key = str(name).strip() + "" + str(opt or "").strip()
        clean = self.rules.get(key) or self._clean_name(name, opt)
        if clean in self.sets:
            out = []
            for c in self.sets[clean]:
                cc = self.canon(c)
                if not self._is_excluded_name(cc):
                    out.append(cc)
            return list(dict.fromkeys(out)), True    # 세트 리뷰
        clean = self.canon(clean)               # 최종 canonical 병합
        fixed = self._foot_fix(clean, name, opt)   # 넥숄더/발/종아리/EMS/프로모 보정
        if fixed is not None:
            return [c for c in fixed if not self._is_excluded_name(c)], False
        if self._is_excluded_name(clean):
            return [], False
        return [clean], False

    def resolve(self, name, opt):
        """→ 마사지기 정리된명 리스트 (사은품/비마사지기 제외, 세트 분해). 빈 리스트=집계제외."""
        return self.resolve_ex(name, opt)[0]


# ── CLI: 설정 생성 + 3개월 분포 리포트 ──
if __name__ == "__main__":
    import csv
    from collections import Counter
    if len(sys.argv) > 1 and sys.argv[1] == "build":
        xlsx = sys.argv[2] if len(sys.argv) > 2 else r"C:\Users\올릿\Downloads\리뷰 대시보드 상품 매핑표.xlsx"
        c = build_config(xlsx)
        print(f"[OK] 설정 생성: {CFG_PATH}  (규칙 {len(c['rules'])} · 별칭 {len(c['aliases'])} · 제외 {len(c['exclude_aliases'])} · 세트 {len(c['set_components'])})")
    rs = ProductResolver()
    print("\n=== 세트 분해 확인 ===")
    for s, comps in rs.sets.items():
        got = rs.resolve(s, "")
        print(f"  {s[:45]:45} → {got}")
    print("\n=== 3개월 상품 분포 (재매핑 결과) ===")
    for m in ["2026-03", "2026-04", "2026-05"]:
        p = ROOT / f"data/raw/슬룸/{m}/reviews.csv"
        rows = list(csv.DictReader(open(p, encoding="utf-8-sig")))
        def col(r, n): return (r.get(n) or "").strip()
        prod = Counter(); excluded = 0; multi = 0; setrev = 0
        for r in rows:
            got, is_set = rs.resolve_ex(col(r, "상품명"), col(r, "상품옵션"))
            if not got:
                excluded += 1; continue
            if is_set:
                setrev += 1
            if len(got) > 1:
                multi += 1
            for g in got:
                prod[g] += 1
        print(f"\n[{m}] 원본 {len(rows)}건 → 제외(사은품) {excluded} · 세트리뷰 {setrev} · 다중귀속 {multi} · 상품 {len(prod)}종")
        for name, c in prod.most_common(12):
            print(f"    {c:5}  {name}")
